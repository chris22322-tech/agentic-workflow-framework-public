#!/usr/bin/env python3
"""Convert per-stage CSVs into multi-tab .xlsx workbooks."""

import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = Path(__file__).parent

# Header styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Example row styling
EXAMPLE_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

# Cell styling
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Stage definitions: output filename -> list of (tab_name, csv_filename)
STAGES = {
    "Stage 1 - Decompose.xlsx": [
        ("Workflow Inventory", "stage1-workflow-inventory.csv"),
        ("Dependencies", "stage1-dependencies.csv"),
    ],
    "Stage 2 - Select.xlsx": [
        ("Scoring Matrix", "stage2-scoring-matrix.csv"),
        ("Selection Decision Record", "stage2-selection-decision-record.csv"),
    ],
    "Stage 3 - Scope.xlsx": [
        ("Workflow Map", "stage3-workflow-map.csv"),
        ("Data Inventory", "stage3-data-inventory.csv"),
        ("Integration Requirements", "stage3-integration-requirements.csv"),
        ("Constraints & Assumptions", "stage3-constraints-and-assumptions.csv"),
    ],
    "Stage 4 - Design.xlsx": [
        ("Scope-to-Node Mapping", "stage4-scope-to-node-mapping.csv"),
        ("State Schema", "stage4-state-schema.csv"),
        ("Node Specifications", "stage4-node-specifications.csv"),
        ("HIL Interaction Design", "stage4-hil-interaction-design.csv"),
        ("Error Handling", "stage4-error-handling.csv"),
    ],
    "Stage 6 - Evaluate.xlsx": [
        ("Test Cases", "stage6-test-cases.csv"),
        ("Evaluation Dimensions", "stage6-evaluation-dimensions.csv"),
        ("Iteration Log", "stage6-iteration-log.csv"),
        ("Graduation Checklist", "stage6-graduation-checklist.csv"),
        ("Efficiency Metrics", "stage6-efficiency-metrics.csv"),
    ],
}


def read_csv(path: Path) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def add_sheet(wb: Workbook, tab_name: str, csv_path: Path):
    rows = read_csv(csv_path)
    if not rows:
        return

    ws = wb.create_sheet(title=tab_name[:31])  # Excel tab name limit

    # Find which rows are examples (non-empty, non-header) vs blank
    header = rows[0]
    num_cols = len(header)

    for row_idx, row in enumerate(rows, start=1):
        # Pad row to match header length
        while len(row) < num_cols:
            row.append("")

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

            if row_idx == 1:
                # Header row
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
            elif any(cell_val.strip() for cell_val in row):
                # Example row (has content)
                cell.fill = EXAMPLE_FILL

    # Auto-width columns (capped at 50 chars)
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for row_idx in range(1, len(rows) + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            # Use first 80 chars for width calc
            max_len = max(max_len, min(len(str(cell_val)), 80))
        width = min(max(max_len * 1.1, 12), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set row heights for content rows
    for row_idx in range(2, len(rows) + 1):
        row_data = rows[row_idx - 1]
        if any(cell_val.strip() for cell_val in row_data):
            # Content row — calculate height based on longest cell
            max_lines = 1
            for val in row_data:
                lines = len(str(val)) // 45 + 1  # rough wrap estimate
                max_lines = max(max_lines, lines)
            ws.row_dimensions[row_idx].height = max(15, min(max_lines * 15, 200))


def main():
    for xlsx_name, tabs in STAGES.items():
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for tab_name, csv_name in tabs:
            csv_path = TEMPLATE_DIR / csv_name
            if csv_path.exists():
                add_sheet(wb, tab_name, csv_path)
                print(f"  Added tab: {tab_name} (from {csv_name})")
            else:
                print(f"  MISSING: {csv_path}")

        out_path = TEMPLATE_DIR / xlsx_name
        wb.save(out_path)
        print(f"Created: {out_path.name}\n")


if __name__ == "__main__":
    main()
